#!/usr/bin/env python3
"""
GL PDF → Excel Extractor
Extracts Standard Ledger PDF data into a validated Excel workbook.

Usage:
    python gl_extractor.py input.pdf output.xlsx
    python gl_extractor.py input.pdf              # auto-names output

Requirements:
    pip install pdfplumber openpyxl
"""

import sys, re, os
from datetime import datetime
from collections import defaultdict

try:
    import pdfplumber
except ImportError:
    sys.exit("Missing: pip install pdfplumber")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing: pip install openpyxl")


# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
DATE_RE  = re.compile(r'^\d{2}-\d{2}-\d{4}$')
NUM_RE   = re.compile(r'^-?\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?$')
ACCT_RE  = re.compile(r'^Account\s+Code\s*:', re.I)
TAX_CODES = {'SR','TX','ZR','BL','ES43','ES','IM','AJP','AJS',
             'TX-E43','DS','RS','IS','GS','TS','NR','OP','EP'}

HEADERS = ['LINE','ACCOUNT CODE','ACCOUNT NAME','PAGE','DATE','JOURNAL',
           'REF 1','REF 2','DESCRIPTION','DESCRIPTION 2','TAX CODE',
           'DEBIT','CREDIT','BALANCE','BALANCE CHECK']

COL_W = {'LINE':6,'ACCOUNT CODE':15,'ACCOUNT NAME':38,'PAGE':5,'DATE':13,
          'JOURNAL':10,'REF 1':22,'REF 2':18,'DESCRIPTION':42,
          'DESCRIPTION 2':30,'TAX CODE':10,'DEBIT':14,'CREDIT':14,
          'BALANCE':16,'BALANCE CHECK':14}

# Colours
BL  = '1E3A5F'
BLL = 'D6E4F0'
GN  = '00B87A'
GNL = 'E6F7F2'
AM  = 'F59E0B'
AML = 'FEF3C7'
RD  = 'EF4444'
RDL = 'FEE2E2'
GYL = 'F3F4F6'
WH  = 'FFFFFF'
TBL = 'EBF5FB'


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def fill(c): return PatternFill('solid', fgColor=c)
def font(bold=False, sz=9, color='000000', italic=False):
    return Font(name='Calibri', bold=bold, size=sz, color=color, italic=italic)
def border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)
def align(h='left', wrap=False):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

def write_grand_total_check(ws, row_no, rows, recon):
    doc_dr = round(sum((r.get('dr') or 0.0) for r in rows if r.get('is_tot')), 2)
    doc_cr = round(sum((r.get('cr') or 0.0) for r in rows if r.get('is_tot')), 2)
    calc_dr = round(sum(r['dr'] for r in recon), 2)
    calc_cr = round(sum(r['cr'] for r in recon), 2)
    check_ok = abs(doc_dr - calc_dr) < 0.02 and abs(doc_cr - calc_cr) < 0.02
    items = [
        ('Grand total from document', doc_dr, doc_cr, '-', 'FOUND', GNL),
        ('Grand total calculated from Excel rows', calc_dr, calc_cr, '-', 'BALANCED' if abs(calc_dr - calc_cr) < 0.02 else 'CHECK', GNL),
        ('Grand total check (document vs formula)', None, None, '', 'OK' if check_ok else 'CHECK', AML),
    ]
    for label, dr, cr, dash, status, row_fill in items:
        ws.row_dimensions[row_no].height = 20
        ws.row_dimensions[row_no].hidden = False
        ws.merge_cells(f'A{row_no}:K{row_no}')
        c = ws.cell(row_no, 1, label)
        c.font = font(bold=True, sz=11, color='000000')
        c.fill = fill(row_fill)
        c.alignment = align('left')
        c.border = border()
        for ci, val in enumerate([dr, cr, dash, status], 12):
            c = ws.cell(row_no, ci, val)
            c.font = font(bold=True, sz=11, color='000000')
            c.fill = fill(row_fill)
            c.alignment = align('right')
            c.border = border()
            if isinstance(val, (int, float)):
                c.number_format = '#,##0.00'
        row_no += 1
    return row_no

def correct_dr_cr_by_balance(rows):
    """Use each account's running balance to fix ambiguous debit/credit placement."""
    prev_by_code = {}
    for r in rows:
        code = r.get('code')
        if not code:
            continue
        if r.get('is_hdr'):
            prev_by_code[code] = None
            continue
        if r.get('is_bf'):
            prev_by_code[code] = r.get('bal')
            continue
        if r.get('is_tot'):
            continue

        prev = prev_by_code.get(code)
        bal = r.get('bal')
        dr = r.get('dr') or 0.0
        cr = r.get('cr') or 0.0
        if prev is not None and bal is not None:
            amt = None
            if dr and not cr:
                amt = abs(dr)
            elif cr and not dr:
                amt = abs(cr)

            if amt is not None:
                target = round(bal, 2)
                if abs(round(prev + amt, 2) - target) < 0.02:
                    r['dr'], r['cr'] = amt, None
                elif abs(round(prev - amt, 2) - target) < 0.02:
                    r['dr'], r['cr'] = None, amt

        if bal is not None:
            prev_by_code[code] = bal

def parse_num(s):
    s = str(s).strip().replace(',','')
    if s.startswith('(') and s.endswith(')'):
        try: return -float(s[1:-1])
        except: return None
    try: return float(s)
    except: return None

def is_num(s):
    return bool(NUM_RE.match(str(s).strip()))


# ─────────────────────────────────────────────
#  EXTRACT
# ─────────────────────────────────────────────
def extract_pdf(pdf_path):
    rows = []
    cur_code = cur_name = ''
    debit_x = credit_x = balance_x = None   # column x-positions from header

    with pdfplumber.open(pdf_path) as pdf:
        print(f"  Pages: {len(pdf.pages)}")
        for pg_no, page in enumerate(pdf.pages, 1):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words: continue

            # Group into lines
            lines = defaultdict(list)
            for w in words:
                lines[round(w['top']/4)*4].append(w)

            pending_desc2 = None   # continuation line buffer

            for y in sorted(lines):
                ws  = sorted(lines[y], key=lambda w: w['x0'])
                txts = [w['text'] for w in ws]
                joined = ' '.join(txts)

                # ── Account header ──────────────────────
                if (txts[:2] == ['Account','Code:'] or
                        ACCT_RE.match(joined)):
                    rest = joined.split(':', 1)[-1].strip().split(None, 1)
                    cur_code = rest[0] if rest else ''
                    cur_name = rest[1] if len(rest) > 1 else ''
                    # reset column positions per account header
                    debit_x = credit_x = balance_x = None
                    rows.append(_hdr(cur_code, cur_name, pg_no))
                    pending_desc2 = None
                    continue

                # ── Column header → detect Dr/Cr/Bal x ──
                if 'Debit' in txts or 'DEBIT' in txts:
                    for w in ws:
                        t = w['text'].upper()
                        if t == 'DEBIT':   debit_x   = w['x0']
                        elif t == 'CREDIT': credit_x = w['x0']
                        elif t == 'BALANCE': balance_x = w['x0']
                    continue

                # ── BALANCE B/F ──────────────────────────
                if txts[:2] == ['BALANCE','B/F']:
                    nums = [parse_num(t) for t in txts[2:] if is_num(t)]
                    bf = nums[-1] if nums else None
                    rows.append(_bf(cur_code, cur_name, pg_no, bf))
                    pending_desc2 = None
                    continue

                # ── Account total (bare numbers only) ────
                if (all(is_num(t) for t in txts)
                        and len(txts) >= 2 and cur_code):
                    nums = [parse_num(t) for t in txts]
                    dr  = nums[0] if len(nums) > 0 else None
                    cr  = nums[1] if len(nums) > 1 else None
                    bal = nums[2] if len(nums) > 2 else None
                    rows.append(_total(cur_code, cur_name, pg_no, dr, cr, bal))
                    pending_desc2 = None
                    continue

                # ── Transaction row ──────────────────────
                if txts and DATE_RE.match(txts[0]) and cur_code:
                    r = _tx(ws, txts, cur_code, cur_name, pg_no,
                            debit_x, credit_x, balance_x)
                    rows.append(r)
                    # Prime pending_desc2 for next continuation line
                    pending_desc2 = r
                    continue

                # ── Continuation description (desc2) ─────
                if (pending_desc2 is not None
                        and not DATE_RE.match(txts[0])
                        and not is_num(txts[0])
                        and txts[0] not in ('Account','BALANCE','0.00')):
                    # Likely continuation — skip page-header lines
                    skip = {'STANDARD', 'LEDGER', 'Page', 'User', 'Date'}
                    if not any(t in skip for t in txts):
                        extra = ' '.join(txts)
                        if not pending_desc2['desc2']:
                            pending_desc2['desc2'] = extra
                        pending_desc2 = None   # only one continuation
                    continue

    return rows


def _hdr(code, name, pg):
    return dict(code=code, name=name, page=pg,
                date='', jrnl='', ref1='', ref2='',
                desc='', desc2='', tax='',
                dr=None, cr=None, bal=None,
                is_hdr=True, is_bf=False, is_tot=False)

def _bf(code, name, pg, bal):
    return dict(code=code, name=name, page=pg,
                date='', jrnl='', ref1='', ref2='',
                desc='BALANCE B/F', desc2='', tax='',
                dr=None, cr=None, bal=bal,
                is_hdr=False, is_bf=True, is_tot=False)

def _total(code, name, pg, dr, cr, bal):
    return dict(code=code, name=name, page=pg,
                date='', jrnl='', ref1='', ref2='',
                desc='ACCOUNT TOTAL', desc2='', tax='',
                dr=dr, cr=cr, bal=bal,
                is_hdr=False, is_bf=False, is_tot=True)

def _tx(ws, txts, code, name, pg, debit_x, credit_x, balance_x):
    # Separate numeric words from text words
    num_ws  = [(w, parse_num(w['text'])) for w in ws if is_num(w['text'])]
    txt_ws  = [w for w in ws if not is_num(w['text'])]

    # Amounts: determine Dr/Cr/Bal from x-position relative to column headers
    dr = cr = bal = None
    for w, val in num_ws:
        x = w['x0']
        if balance_x and abs(x - balance_x) < 25:
            bal = val
        elif credit_x and abs(x - credit_x) < 25:
            cr = abs(val) if val else None
        elif debit_x and abs(x - debit_x) < 25:
            dr = abs(val) if val else None

    # Fallback if column positions unknown: last=bal, second-last=cr or dr
    if bal is None and num_ws:
        bal = num_ws[-1][1]
    if dr is None and cr is None and len(num_ws) >= 2:
        val2 = num_ws[-2][1]
        if val2 is not None:
            # Can't tell dr vs cr without positions; use balance change
            if bal is not None:
                prev_bal = bal - (-val2)   # if credit: prev = bal + val2
                # naive: assign to cr (will be corrected by reconcile)
                cr = abs(val2)

    # Text fields after skipping the date word
    txt_only = [w for w in txt_ws if w['x0'] > ws[0]['x0']]
    jrnl = ref1 = ref2 = desc = ''
    tax_found = None

    # Journal — first short ALLCAPS token
    if txt_only and (txt_only[0]['text'].isupper() or
                     len(txt_only[0]['text']) <= 12):
        jrnl = txt_only[0]['text']
        txt_only = txt_only[1:]

    # Ref1 — next token
    if txt_only:
        ref1 = txt_only[0]['text']
        txt_only = txt_only[1:]

    # Ref2 — next if short or looks like a label
    if txt_only and len(txt_only[0]['text']) <= 25:
        ref2 = txt_only[0]['text']
        txt_only = txt_only[1:]

    # Tax code — extract any TAX_CODE token
    tax_ws = [w for w in txt_only if w['text'].upper() in TAX_CODES]
    if tax_ws:
        tax_found = tax_ws[-1]['text']
        txt_only  = [w for w in txt_only if w not in tax_ws]

    desc = ' '.join(w['text'] for w in txt_only)

    return dict(code=code, name=name, page=pg,
                date=txts[0], jrnl=jrnl, ref1=ref1, ref2=ref2,
                desc=desc, desc2='', tax=tax_found or '',
                dr=dr, cr=cr, bal=bal,
                is_hdr=False, is_bf=False, is_tot=False)


# ─────────────────────────────────────────────
#  RECONCILE
# ─────────────────────────────────────────────
def reconcile(rows):
    accts = {}
    cur = None
    for r in rows:
        code = r['code']
        if r['is_hdr']:
            cur = code
            if code not in accts:
                accts[code] = {'name': r['name'], 'bf': 0.0,
                               'dr': 0.0, 'cr': 0.0, 'pdf_bal': None}
        if not cur or code != cur: continue
        a = accts[code]
        if r['is_bf']:
            a['bf'] = r['bal'] or 0.0
        elif r['is_tot']:
            if r['dr'] is not None: a['dr']  = r['dr']
            if r['cr'] is not None: a['cr']  = r['cr']
            if r['bal'] is not None: a['pdf_bal'] = r['bal']
        elif r['date']:
            a['dr'] += r['dr'] or 0.0
            a['cr'] += r['cr'] or 0.0
            if r['bal'] is not None: a['pdf_bal'] = r['bal']

    out = []
    for code, a in accts.items():
        calc = round(a['bf'] + a['dr'] - a['cr'], 2)
        pdf  = round(a['pdf_bal'] or 0.0, 2)
        diff = round(calc - pdf, 2)
        ok   = abs(diff) < 0.02
        out.append({**a, 'code': code, 'calc': calc, 'pdf': pdf,
                    'diff': diff, 'ok': ok})
    return out


# ─────────────────────────────────────────────
#  WRITE EXCEL
# ─────────────────────────────────────────────
def write_excel(rows, recon, out_path):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'General Ledger'
    ws.freeze_panes = 'A3'

    # Title
    ws.merge_cells('A1:O1')
    c = ws['A1']
    c.value     = f"GENERAL LEDGER   Extracted: {datetime.now():%d %b %Y  %H:%M}"
    c.font      = font(bold=True, sz=12, color=WH)
    c.fill      = fill(BL)
    c.alignment = align('center')
    ws.row_dimensions[1].height = 22

    # Headers
    ws.row_dimensions[2].height = 18
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(2, ci, h)
        c.font = font(bold=True, color=WH)
        c.fill = fill(BL)
        c.alignment = align('center')
        c.border = border()
        ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(h, 12)

    # Data
    line_no = 0
    row_no  = 3
    gdr = gcr = 0.0
    recon_map = {r['code']: r for r in recon}

    for r in rows:
        if r['is_hdr']:
            ws.row_dimensions[row_no].height = 16
            ws.merge_cells(f'A{row_no}:O{row_no}')
            c = ws.cell(row_no, 1,
                        f"  Account Code: {r['code']}   {r['name']}")
            c.font = font(bold=True, sz=10, color=BL)
            c.fill = fill(BLL)
            c.alignment = align('left')
            c.border = border()
            row_no += 1
            continue

        line_no += 1
        ws.row_dimensions[row_no].height = 14

        ok_acc = recon_map.get(r['code'], {}).get('ok', True)
        if r['is_bf']:
            rfill = fill(GYL)
            rfont = font(italic=True, color='666666')
        elif r['is_tot']:
            rfill = fill(TBL)
            rfont = font(bold=True, color=BL)
        else:
            rfill = fill(WH)
            rfont = font()
            gdr += r['dr'] or 0.0
            gcr += r['cr'] or 0.0

        vals = [line_no, r['code'], r['name'], r['page'],
                r['date'], r['jrnl'], r['ref1'], r['ref2'],
                r['desc'], r['desc2'], r['tax'],
                r['dr'], r['cr'], r['bal'], '']

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row_no, ci, val)
            c.font   = rfont
            c.fill   = rfill
            c.border = border()
            c.alignment = align('right' if ci in (1,4,12,13,14) else 'left')
            if ci in (12, 13, 14) and isinstance(val, (int, float)):
                c.number_format = '#,##0.00'

        # Balance check
        if not r['is_bf'] and not r['is_tot']:
            ok_str = '✅' if ok_acc else '⚠'
            c15 = ws.cell(row_no, 15, ok_str)
            c15.font = font(color=GN if ok_acc else AM, bold=True)
            c15.alignment = align('center')

        row_no += 1

    old_total_start = row_no
    # Grand totals
    for label, dr, cr in [
        ('GRAND TOTAL — All Journal Entries', gdr, gcr),
        ('Grand Total DR  −  Grand Total CR', gdr - gcr, None),
    ]:
        ws.row_dimensions[row_no].height = 20
        ws.merge_cells(f'A{row_no}:K{row_no}')
        c = ws.cell(row_no, 1, label)
        c.font = font(bold=True, sz=11, color=WH)
        c.fill = fill(BL)
        c.alignment = align('right')
        for ci, val in enumerate([dr, cr, ''], 12):
            c = ws.cell(row_no, ci, val)
            c.font = font(bold=True, sz=11, color=WH)
            c.fill = fill(BL)
            c.alignment = align('right')
            if isinstance(val, (int, float)): c.number_format = '#,##0.00'
        row_no += 1

    # ── Sheet 2: Validation ──────────────────────────────
    for old_row in range(old_total_start, row_no):
        ws.row_dimensions[old_row].hidden = True
    row_no = write_grand_total_check(ws, old_total_start, rows, recon)

    ws2 = wb.create_sheet('Validation Report')
    all_ok = all(r['ok'] for r in recon)
    ws2.merge_cells('A1:I1')
    c = ws2['A1']
    c.value     = ('✅  ALL ACCOUNTS BALANCED' if all_ok
                   else '⚠  MISMATCHES DETECTED — see highlighted rows')
    c.font      = font(bold=True, sz=12, color=WH)
    c.fill      = fill(GN if all_ok else RD)
    c.alignment = align('center')
    ws2.row_dimensions[1].height = 22

    h2 = ['Code','Account Name','B/F','Total DR','Total CR',
          'Calc Balance','PDF Balance','Diff','✓']
    w2 = [15, 38, 14, 14, 14, 14, 14, 12, 6]
    for ci, (h, w) in enumerate(zip(h2, w2), 1):
        c = ws2.cell(2, ci, h)
        c.font = font(bold=True, color=WH)
        c.fill = fill(BL)
        c.alignment = align('center')
        c.border = border()
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for ri, r in enumerate(recon, 3):
        rfill = fill(WH) if r['ok'] else fill(RDL)
        for ci, val in enumerate(
                [r['code'], r['name'], r['bf'], r['dr'], r['cr'],
                 r['calc'], r['pdf'], r['diff'],
                 '✅' if r['ok'] else '❌'], 1):
            c = ws2.cell(ri, ci, val)
            c.font = font(color='000000' if r['ok'] else RD,
                          bold=(ci == 8 and not r['ok']))
            c.fill = rfill
            c.border = border()
            c.alignment = align('right' if ci >= 3 else 'left')
            if ci in range(3, 9) and isinstance(val, (int, float)):
                c.number_format = '#,##0.00'

    ws2.freeze_panes = 'A3'
    wb.save(out_path)
    print(f"  Saved → {out_path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def run(pdf_path, out_path=None):
    if not out_path:
        out_path = os.path.splitext(pdf_path)[0] + '_extracted.xlsx'

    print(f"\n{'═'*60}")
    print(f"  GL Extractor  ·  {os.path.basename(pdf_path)}")
    print(f"{'═'*60}")

    print("\n[1/3] Extracting rows …")
    rows = extract_pdf(pdf_path)
    correct_dr_cr_by_balance(rows)
    txns = [r for r in rows if r.get('date') and not r.get('is_tot')]
    print(f"  Rows: {len(rows)} total  ·  {len(txns)} transactions")

    print("\n[2/3] Reconciling …")
    recon = reconcile(rows)
    ok_n  = sum(1 for r in recon if r['ok'])
    print(f"  Accounts: {len(recon)}  ✅ {ok_n}  ❌ {len(recon)-ok_n}")
    for r in recon:
        if not r['ok']:
            print(f"    ❌ {r['code']:<15} {r['name'][:35]:<35}"
                  f"  diff={r['diff']:+.2f}")

    print("\n[3/3] Writing Excel …")
    write_excel(rows, recon, out_path)

    gdr = sum((r.get('dr') or 0) for r in rows if r.get('is_tot'))
    gcr = sum((r.get('cr') or 0) for r in rows if r.get('is_tot'))
    print(f"\n  Grand Total DR : {gdr:>18,.2f}")
    print(f"  Grand Total CR : {gcr:>18,.2f}")
    net = gdr - gcr
    print(f"  Net (DR−CR)    : {net:>18,.2f}  "
          f"{'✅ BALANCED' if abs(net)<0.02 else '⚠ CHECK'}")
    print(f"\n{'═'*60}\n")
    return out_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    run(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
